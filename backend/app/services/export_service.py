"""Data export service for generating CSV, PDF, and Excel exports."""

import io
import csv
from datetime import datetime, date
from typing import List, Optional, BinaryIO, Union
from decimal import Decimal
from uuid import UUID
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.db.models.data import Expense, Budget, Loan, Goal
from app.core.logging import get_logger

logger = get_logger(__name__)


class ExportService:
    """Service for exporting data in various formats."""

    def __init__(self, db_session: AsyncSession):
        """Initialize export service."""
        self.db_session = db_session

    def _convert_uuid(self, user_id: Union[str, UUID]) -> UUID:
        """Convert user_id to UUID if it's a string."""
        if isinstance(user_id, str):
            return UUID(user_id)
        return user_id

    async def export_expenses_csv(
        self,
        user_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> io.BytesIO:
        """
        Export expenses to CSV format.

        Args:
            user_id: User UUID
            start_date: Optional start date filter
            end_date: Optional end date filter

        Returns:
            BytesIO object containing CSV data
        """
        # Convert user_id to UUID
        user_uuid = self._convert_uuid(user_id)
        
        # Fetch expenses
        query = select(Expense).where(Expense.user_id == user_uuid)

        if start_date:
            query = query.where(Expense.date >= start_date)
        if end_date:
            query = query.where(Expense.date <= end_date)

        result = await self.db_session.execute(query.order_by(Expense.date.desc()))
        expenses = result.scalars().all()

        # Create CSV
        output = io.BytesIO()
        text_wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)

        writer = csv.DictWriter(
            text_wrapper,
            fieldnames=[
                "Date",
                "Category",
                "Subcategory",
                "Amount",
                "Merchant",
                "Payment Method",
                "Description",
            ],
        )

        writer.writeheader()
        for expense in expenses:
            writer.writerow(
                {
                    "Date": expense.date.isoformat(),
                    "Category": expense.category,
                    "Subcategory": expense.subcategory or "",
                    "Amount": f"₹{float(expense.amount):,.2f}",
                    "Merchant": expense.merchant or "",
                    "Payment Method": expense.payment_method or "",
                    "Description": expense.description or "",
                }
            )

        text_wrapper.flush()
        text_wrapper.detach()  # Detach to prevent closing the BytesIO
        output.seek(0)

        logger.info(
            f"Expenses exported to CSV",
            user_id=user_id,
            count=len(expenses),
        )

        return output

    async def export_budgets_csv(
        self,
        user_id: str,
        month: Optional[date] = None,
    ) -> io.BytesIO:
        """
        Export budgets to CSV format.

        Args:
            user_id: User UUID
            month: Optional month filter

        Returns:
            BytesIO object containing CSV data
        """
        # Convert user_id to UUID
        user_uuid = self._convert_uuid(user_id)
        
        # Fetch budgets
        query = select(Budget).where(Budget.user_id == user_uuid)

        if month:
            # Filter by month
            query = query.where(Budget.month >= month)

        result = await self.db_session.execute(query.order_by(Budget.month.desc()))
        budgets = result.scalars().all()

        # Create CSV
        output = io.BytesIO()
        text_wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)

        writer = csv.DictWriter(
            text_wrapper,
            fieldnames=[
                "Month",
                "Category",
                "Allocated Amount",
                "Spent Amount",
                "Remaining Amount",
                "Utilization %",
            ],
        )

        writer.writeheader()
        for budget in budgets:
            remaining = float(budget.allocated_amount) - float(budget.spent_amount)
            utilization = (
                (float(budget.spent_amount) / float(budget.allocated_amount) * 100)
                if budget.allocated_amount
                else 0
            )

            writer.writerow(
                {
                    "Month": budget.month.isoformat(),
                    "Category": budget.category,
                    "Allocated Amount": f"₹{float(budget.allocated_amount):,.2f}",
                    "Spent Amount": f"₹{float(budget.spent_amount):,.2f}",
                    "Remaining Amount": f"₹{remaining:,.2f}",
                    "Utilization %": f"{utilization:.1f}%",
                }
            )

        text_wrapper.flush()
        text_wrapper.detach()  # Detach to prevent closing the BytesIO
        output.seek(0)

        logger.info(
            f"Budgets exported to CSV",
            user_id=user_id,
            count=len(budgets),
        )

        return output

    async def export_loans_csv(self, user_id: str) -> io.BytesIO:
        """
        Export loans to CSV format.

        Args:
            user_id: User UUID

        Returns:
            BytesIO object containing CSV data
        """
        # Convert user_id to UUID
        user_uuid = self._convert_uuid(user_id)
        
        # Fetch loans
        result = await self.db_session.execute(
            select(Loan).where(Loan.user_id == user_uuid).order_by(Loan.start_date.desc())
        )
        loans = result.scalars().all()

        # Create CSV
        output = io.BytesIO()
        text_wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)

        writer = csv.DictWriter(
            text_wrapper,
            fieldnames=[
                "Loan Type",
                "Lender",
                "Principal Amount",
                "Interest Rate %",
                "EMI Amount",
                "Outstanding Balance",
                "Status",
                "Start Date",
                "Next Due Date",
                "Remaining Months",
            ],
        )

        writer.writeheader()
        for loan in loans:
            writer.writerow(
                {
                    "Loan Type": loan.loan_type,
                    "Lender": loan.lender_name or "",
                    "Principal Amount": f"₹{float(loan.principal_amount):,.2f}",
                    "Interest Rate %": f"{float(loan.interest_rate):.2f}%",
                    "EMI Amount": f"₹{float(loan.emi_amount):,.2f}",
                    "Outstanding Balance": f"₹{float(loan.outstanding_balance):,.2f}",
                    "Status": loan.status,
                    "Start Date": loan.start_date.isoformat(),
                    "Next Due Date": loan.next_due_date.isoformat(),
                    "Remaining Months": loan.remaining_months,
                }
            )

        text_wrapper.flush()
        text_wrapper.detach()  # Detach to prevent closing the BytesIO
        output.seek(0)

        logger.info(
            f"Loans exported to CSV",
            user_id=user_id,
            count=len(loans),
        )

        return output

    async def export_goals_csv(self, user_id: str) -> io.BytesIO:
        """
        Export goals to CSV format.

        Args:
            user_id: User UUID

        Returns:
            BytesIO object containing CSV data
        """
        # Convert user_id to UUID
        user_uuid = self._convert_uuid(user_id)
        
        # Fetch goals
        result = await self.db_session.execute(
            select(Goal).where(Goal.user_id == user_uuid).order_by(Goal.created_at.desc())
        )
        goals = result.scalars().all()

        # Create CSV
        output = io.BytesIO()
        text_wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)

        writer = csv.DictWriter(
            text_wrapper,
            fieldnames=[
                "Goal Name",
                "Type",
                "Target Amount",
                "Current Amount",
                "Progress %",
                "Target Date",
                "Status",
                "Priority",
                "Days Remaining",
            ],
        )

        writer.writeheader()
        for goal in goals:
            progress = (
                (float(goal.current_amount) / float(goal.target_amount) * 100)
                if goal.target_amount
                else 0
            )
            days_remaining = (goal.target_date - date.today()).days

            writer.writerow(
                {
                    "Goal Name": goal.goal_name,
                    "Type": goal.goal_type,
                    "Target Amount": f"₹{float(goal.target_amount):,.2f}",
                    "Current Amount": f"₹{float(goal.current_amount):,.2f}",
                    "Progress %": f"{progress:.1f}%",
                    "Target Date": goal.target_date.isoformat(),
                    "Status": goal.status,
                    "Priority": goal.priority,
                    "Days Remaining": days_remaining,
                }
            )

        text_wrapper.flush()
        text_wrapper.detach()  # Detach to prevent closing the BytesIO
        output.seek(0)

        logger.info(
            f"Goals exported to CSV",
            user_id=user_id,
            count=len(goals),
        )

        return output

    async def export_expenses_excel(
        self,
        user_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> io.BytesIO:
        """
        Export expenses to Excel format.

        Args:
            user_id: User UUID
            start_date: Optional start date filter
            end_date: Optional end date filter

        Returns:
            BytesIO object containing Excel data
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        # Convert user_id to UUID
        user_uuid = self._convert_uuid(user_id)

        # Fetch expenses
        query = select(Expense).where(Expense.user_id == user_uuid)

        if start_date:
            query = query.where(Expense.date >= start_date)
        if end_date:
            query = query.where(Expense.date <= end_date)

        result = await self.db_session.execute(query.order_by(Expense.date.desc()))
        expenses = result.scalars().all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Add headers
        headers = [
            "Date",
            "Category",
            "Subcategory",
            "Amount",
            "Merchant",
            "Payment Method",
            "Description",
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for expense in expenses:
            ws.append(
                [
                    expense.date.isoformat(),
                    expense.category,
                    expense.subcategory or "",
                    f"₹{float(expense.amount):,.2f}",
                    expense.merchant or "",
                    expense.payment_method or "",
                    expense.description or "",
                ]
            )

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 25

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            f"Expenses exported to Excel",
            user_id=user_id,
            count=len(expenses),
        )

        return output

    async def export_complete_financial_data_excel(
        self,
        user_id: str,
    ) -> io.BytesIO:
        """
        Export all financial data to a single Excel file with multiple sheets.

        Args:
            user_id: User UUID

        Returns:
            BytesIO object containing Excel data
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        # Convert user_id to UUID
        user_uuid = self._convert_uuid(user_id)

        # Fetch all data
        expenses_result = await self.db_session.execute(
            select(Expense).where(Expense.user_id == user_uuid).order_by(Expense.date.desc())
        )
        expenses = expenses_result.scalars().all()

        budgets_result = await self.db_session.execute(
            select(Budget).where(Budget.user_id == user_uuid).order_by(Budget.month.desc())
        )
        budgets = budgets_result.scalars().all()

        loans_result = await self.db_session.execute(
            select(Loan).where(Loan.user_id == user_uuid).order_by(Loan.start_date.desc())
        )
        loans = loans_result.scalars().all()

        goals_result = await self.db_session.execute(
            select(Goal).where(Goal.user_id == user_uuid).order_by(Goal.created_at.desc())
        )
        goals = goals_result.scalars().all()

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Create Expenses sheet
        if expenses:
            ws = wb.create_sheet("Expenses")
            headers = ["Date", "Category", "Subcategory", "Amount", "Merchant", "Payment Method", "Description"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for expense in expenses:
                ws.append([
                    expense.date.isoformat(),
                    expense.category,
                    expense.subcategory or "",
                    f"₹{float(expense.amount):,.2f}",
                    expense.merchant or "",
                    expense.payment_method or "",
                    expense.description or "",
                ])

        # Create Budgets sheet
        if budgets:
            ws = wb.create_sheet("Budgets")
            headers = ["Month", "Category", "Allocated Amount", "Spent Amount", "Remaining Amount", "Utilization %"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for budget in budgets:
                remaining = float(budget.allocated_amount) - float(budget.spent_amount)
                utilization = (float(budget.spent_amount) / float(budget.allocated_amount) * 100) if budget.allocated_amount else 0
                ws.append([
                    budget.month.isoformat(),
                    budget.category,
                    f"₹{float(budget.allocated_amount):,.2f}",
                    f"₹{float(budget.spent_amount):,.2f}",
                    f"₹{remaining:,.2f}",
                    f"{utilization:.1f}%",
                ])

        # Create Loans sheet
        if loans:
            ws = wb.create_sheet("Loans")
            headers = ["Loan Type", "Lender", "Principal Amount", "Interest Rate %", "EMI Amount", "Outstanding Balance", "Status"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for loan in loans:
                ws.append([
                    loan.loan_type,
                    loan.lender_name or "",
                    f"₹{float(loan.principal_amount):,.2f}",
                    f"{float(loan.interest_rate):.2f}%",
                    f"₹{float(loan.emi_amount):,.2f}",
                    f"₹{float(loan.outstanding_balance):,.2f}",
                    loan.status,
                ])

        # Create Goals sheet
        if goals:
            ws = wb.create_sheet("Goals")
            headers = ["Goal Name", "Type", "Target Amount", "Current Amount", "Progress %", "Target Date", "Status", "Priority"]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for goal in goals:
                progress = (float(goal.current_amount) / float(goal.target_amount) * 100) if goal.target_amount else 0
                ws.append([
                    goal.goal_name,
                    goal.goal_type,
                    f"₹{float(goal.target_amount):,.2f}",
                    f"₹{float(goal.current_amount):,.2f}",
                    f"{progress:.1f}%",
                    goal.target_date.isoformat(),
                    goal.status,
                    goal.priority,
                ])

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(
            f"Complete financial data exported to Excel",
            user_id=user_id,
            sheets=wb.sheetnames,
        )

        return output
